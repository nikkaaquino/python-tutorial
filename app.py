#Intro
'''
print("Nikka Aquino")
print('o----')
print(' ||||')
print('*' * 10)
'''

#Variables
'''
price = 10
rating = 4.9
name = 'Mosh'
is_published = False
print(price)
'''

#Exercise #1
'''
full_name = 'John Smith'
age = 20
is_new = True
'''

#Receiving Input
'''
name = input('What is your name? ')
print('Hi ' + name)
'''

#Exercise #2
'''
name = input('What is your name? ')
color = input('What is your favorite color? ')
print(name + " likes " + color)
'''

#Type Conversion
'''
birth_year = input('Birth year: ')
print(type(birth_year))
age = 2024 - int(birth_year)
print(type(age))
print(age)
'''
#Exercise #3
'''
pounds = input('Your weight in pounds: ')
kilograms = int(pounds) * 0.453592
print(kilograms)
'''

#Strings
'''
course = "Python's Course for Beginners"
python = 'Pythons Course for "Beginners"'
print(course)
print(python)
'''


# course = ''' 
# Hi John,

# This is our first email to you.

# Thank you.

# '''
# print(course)


#String with square brackets
'''
course = 'Python for Beginners'
print(course[0])
print(course[-1])
print(course[0:3])
print(course[0:])
print(course[1:])
print(course[:5])
print(course[:])

name = 'Jennifer'
print(name[1:-1])
'''

#Formatted Strings
'''
first = 'John'
last = 'Smith'
message = first + ' [' + last + '] is a coder.'
msg = f'{first} [{last}] is a coder'
print(msg)
'''

#String Methods
'''
course = 'Python for Beginners'
print(len(course))
print(course.upper())
print(course)
print(course.lower())
print(course.find('P'))
print(course.find('o'))
print(course.find('Beginners'))
print(course.replace('Beginners', 'Absolute Beginners'))
print('Python' in course)
print('python' in course)
print(course.title())
'''

#Arithmetic Operations
'''
print(10 / 3)
print(10 // 3) #int
print(10 % 3) #modulo
print(10 ** 3) #exponent
x = 10
# x = x + 10
# x += 3 #augmented assignment operator
x -= 3
print(x)
'''

#Operator Precedence
'''
x = (10 + 3) * 2 ** 2
print(x)

#parenthesis
# exponentiation 2 ** 3
# multiplication or division
#addition or subtruction
'''

#Math Functions
'''
x = 2.9
print(round(x))
print(abs(-2.9))

import math
print(math.ceil(2.9))
print(math.floor(2.9))
'''

#If Statements
'''
if it's a hot
    It' a hot day
    Drink plenty of water
otherwise if it's cold
    It's a cold day
    Wear warm clothes
otherwise
    It's a lovely day
'''
'''
is_hot = False
is_cold = True

if is_hot:
    print("It' a hot day")
    print("Drink plenty of water")
elif is_cold:
    print("It's a cold day")
    print("Wear warm clothes")
else:
    print("It's a lovely day")
print("Enjoy your day")
'''
#Exercise #4
'''
house_price = 1_000_000
is_good_credit = True

if is_good_credit:
   down_payment = house_price * 0.10
else:
    down_payment = house_price * 0.20
print(f"Down payment: ${down_payment}")
'''

#Logical Operators
'''
#AND:
#OR
#NOT

has_high_income = True
#has_good_credit = True
has_criminal_record = True

#if has_high_income and has_good_credit:
#if has_high_income or has_good_credit:
if has_high_income and not has_criminal_record:
    print("Eligble for loan")
'''

#Comparison Operators
'''
temperature = 35

if temperature != 30:
    print("It's a hot day")
else:
    print("It's not a hot day")
'''
#Exercise #5
'''
name = "John Smith"

if len(name) < 3:
    print("Name must be atleast 3 characters")
elif len(name) > 50:
    print("Name can be a maximum of 50 characters")
else:
    print("Name looks good")
'''

#Weight Converter Program (Simple Project)
'''
weight = int(input("Weight: "))
unit = input("(L)bs or (K)g: ")

if unit.upper() == "L":
    weight_converted = weight * 0.45
    print(f'You are {weight_converted} pounds')
else:
     weight_converted = weight/0.45
     print(f'You are {weight_converted} kilograms')
'''
#While Loops
'''
i = 1
while i <= 5:
    print('*' * i)
    i = i + 1
print("Done")
'''

#Guessing Game
'''
secret_number = 9
guess_count = 0
guess_limit = 3

while guess_count < guess_limit:
    guess = int(input('Guess: '))
    guess_count += 1
    if guess == secret_number:
        print("You won!")
        break
else:
    print("Sorry you failed.")
'''

#Building the Car Game
'''
command = ""
started = False

while True:
    command = input("> ").lower()
    if command == "start":
        if started:
            print("Car is already started!")
        else:
            started = True
            print("Car started...")
    elif command == "stop":
        if not started:
            print("Car is already stopped!")
        else:
            started = False
            print("Car stopped.")
    elif command == "help":
        print("""
start - to start the car
stop - to stop the car
quit - to quit
        """)
    elif command == "quit":
        break
    else:
        print("Sorry, I don't understand that")
'''

#For Loops
'''
for item in 'Python':
    print(item)
'''
'''
for item in ['Mosh', 'John', 'Sarah']:
    print(item)
'''
'''
for item in [1, 2, 3, 4, 5]:
    print(item)
'''
'''
#for item in range(10):
#for item in range(5, 10): #start and end
for item in range(5, 10, 2): #two steps forward
    print(item)
'''
'''
prices = [10, 20, 30]
total = 0

for price in prices:
   total += price
print(f"Total: {total}")
'''

#Nested Loops
'''
for x in range(4):
    for y in range(3):
        print(f'{x},{y}')
'''

#Challenge
'''
numbers = [5, 2, 5, 2, 2]

for x_count in numbers:
    output = ''
    for count in range(x_count):
        output += 'x'
    print(output)
    #print('x' * x)
   
'''

#List
'''
names = ['John', 'Bob', 'Mosh', 'Sarah', 'Mary']
print(names[2:4])
'''

#Find the largest number
'''
numbers = [3, 6, 2, 8, 4, 10]
max = numbers[0]

for number in numbers:
    if number > max:
        max = number
print(max)
'''

#2D Lists
'''
[
    1 2 3
    4 5 6
    7 8 9
]

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]

#print(matrix[0][1])
#matrix[0][1] = 20
#print(matrix[0][1])

for row in matrix:
    for item in row:
        print(item)
'''

#List Methods
'''
numbers = [5, 2, 1, 7, 4]
#numbers.insert(0, 10)
#numbers.remove(5)
#numbers.clear()
#numbers.pop() //remove the last value from the list
#print(numbers)
#print(numbers.index(5))
#print(50 in numbers) #false - occurunces in list
#print(numbers.count(5))
#numbers.sort()
#numbers.reverse()

numbers2 = numbers.copy()
numbers.append(10)
print(numbers)
print(numbers2)
'''

#Exercise
'''
numbers = [1,2,1,2,3,4,5]
uniques = []

for number in numbers:
    if number not in uniques:
        uniques.append(number)
print(uniques)
'''

#Tupples
'''
numbers = (1,2,3)
print(numbers[0])
'''

#Unpacking
'''
coordinates = (1, 2, 3)
#x = coordinates[1]
#y = coordinates[1]
#z = coordinates[2]

x, y, z = coordinates
print(x)
'''

#Dictionaries
'''
customer = {
    "name": "John Smith",
    "age" : 30,
    "is_verified": True
}

#print(customer["name"])
customer["name"] = "Jack Smith"
print(customer.get("birthdate", "Jan 1 1980"))
print(customer["name"])

'''

#Exercise
'''
phone = input("Phone: ")
digits_mapping = {
    "1" : "One",
    "2" : "Two",
    "3" : "Three",
    "4" : "Four"
}

output = ""

for ch in phone:
    output+= digits_mapping.get(ch, "!") + " "
print(output)
'''

#Emoji Converter
'''
message = input(">")
words = message.split(' ')
emojis = {
    ":)" : "🙂",
    ":(" : "😔"
}

output = ""
for word in words:
    output += emojis.get(word, word) + " "
print(output)
'''

#Functions
'''
def greet_user():
    print('Hi there!')
    print('Welcome aboard')

print("Start")
greet_user()
print("Finish")

'''

#Parameters
'''
def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard')

print("Start")
greet_user("John", "Smith")
#greet_user("Mary")
print("Finish")
'''

#Keyword Arguments
'''
def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard')

print("Start")
greet_user(last_name="Smith", first_name="John")
#calc_cost(total=50, shipping=5, discount=0.1)
print("Finish")
'''

#Return Statement
'''
def square(number):
    return number * number

#result = square(3)
#print(result)
print(square(3))
'''

#Creating a reusable function
'''
def emojis_conv(message):
    words = message.split(' ')
    emojis = {
        ":)" : "🙂",
        ":(" : "😔"
    }

    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return(output)

message = input(">")
print(emojis_conv(message))
'''
#Exceptions
'''
try:
    age = int(input('Age: '))
    income = 20_000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print('Age cannot be 0.')
except ValueError:
    print('Invalid value')
'''

#Classes
'''
class Point:
    def move(self):
        print("move")

    def draw(self):
        print("draw")

point1 = Point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()

point2 = Point()
point2.x = 1
print(point2.x)

'''

#Constructors
'''
class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y
    def move(self):
        print("move")

    def draw(self):
        print("draw")

point = Point(10, 20)
point.x = 11
print(point.x)
'''

#Exercise
'''
class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"Hi, I am {self.name}")

person = Person(input(">"))
person.talk()
'''

#Inheritance
'''
class Mammal: 
    def walk(self):
        print("walk")

class Dog(Mammal):
    #pass
    def bark(self):
        print("bark")   

class Cat(Mammal):
    #pass
    def be_annoying(self):
        print("annoying")

dog1 =  Dog()
dog1.walk()
dog1.bark()

cat1 = Cat()
cat1.be_annoying()
'''

#Modules
'''
import converters
from converters import kg_to_lbs

kg_to_lbs()

print(converters.kg_to_lbs(70))
'''

#Exercise
'''
from utils import find_max

numbers = [10, 3, 6, 2]
maximum = find_max(numbers)
print(maximum)
#print(max(numbers))
'''

#Packages
'''
#import ecommerce.shipping
#ecommerce.shipping.calc_shipping()

#from ecommerce.shipping import calc_shipping
#calc_shipping()

from ecommerce import shipping
shipping.calc_shipping()
'''

#Generating Random Values
'''
import random

#for i in range(3):
   #print(random.random())
   #print(random.randint(10, 20))

members = ['John', 'Mary', 'Bob', 'Mosh']
leader = random.choice(members)
print(leader)
'''

#Exercise
'''
import random

class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second

dice = Dice()
print(dice.roll())
'''

#Working with Directories
'''
from pathlib import Path

#Absolute Path
# c:\Program Files\Microsoft
#/usr/local/bin - linux
#Relative Path

path = Path()

for file in path.glob('*.py'):
    print(file)

#print(path.rmdir()) #remove directory
#print(path.mkdir())
#print(path.exists())
'''

#Pypi and Pip
#pip install openpyxl

# Automation with Python

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    #cell = sheet['a1']
    #sheet.cell(1, 1)

    #print(cell.value)
    #print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9

        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value =corrected_price

    values = Reference(sheet, 
            min_row = 2, 
            max_row = sheet.max_row,
            min_col = 4,
            max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
        #print(cell.value)
        #print(row)

